import json
from openpyxl import load_workbook

PNL = r"c:\Users\shaad\Downloads\pnl-BK2660 (2).xlsx"
wb = load_workbook(PNL, data_only=True)
ws = wb["F&O"]
rows = []
for row in ws.iter_rows(values_only=True):
    rows.append(row)

print("=== F&O sheet full dump ===")
for i, row in enumerate(rows):
    if any(c is not None and str(c).strip() for c in row):
        print(i, row)
